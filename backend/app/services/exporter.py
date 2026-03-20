from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.models.part_models import ChangeListMatch, ReleaseNoteRow
from app.services.excel_testplan_builder import (
    build_no_inventory_table,
    build_release_note_table,
    build_system_fw_table,
    build_testplan_rows,
    build_version_note_table,
    infer_os_order,
)
from app.services.os_rule_engine import TestPlanRowState
from app.services.selection_engine import NoInventoryPart


def generate_output_filename(project_name: str, date_str: Optional[str] = None) -> str:
    if not date_str:
        date_str = datetime.now().strftime("%Y%m%d")
    return f"TestPlan1_{project_name}_{date_str}.xlsx"


class TestPlanExporter:
    """
    导出器修正规则：
    1. HDD / SSD：仅 FW Result 保留，Driver Result 强制清空
    2. BIOS/BMC/FPGA/DDlist 版本备注类列：统一强制留空
    3. 其余类型不改现有 builder 生成的 FW/Driver 逻辑，仅清空备注列
    """

    def __init__(self, output_dir: str | Path = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_testplan_excel(
        self,
        project_name: str,
        row_states: Sequence[TestPlanRowState],
        no_inventory_parts: Sequence[NoInventoryPart],
        release_rows: Optional[Sequence[ReleaseNoteRow]] = None,
        system_fw_rows: Optional[Sequence[Dict[str, object]]] = None,
        version_note_lines: Optional[Sequence[str]] = None,
        date_str: Optional[str] = None,
    ) -> Path:
        filename = generate_output_filename(project_name, date_str)
        output_path = self.output_dir / filename

        wb = Workbook()
        main_ws = wb.active
        main_ws.title = "DDlist"

        os_order = infer_os_order(row_states)
        headers, rows = build_testplan_rows(row_states, os_order=os_order)

        # 关键：对 builder 生成的结果做最终兜底清洗
        rows = self._sanitize_testplan_rows(headers, rows)

        self._write_sheet(main_ws, headers, rows)

        release_ws = wb.create_sheet(title="Release note")
        self._write_sheet(release_ws, *build_release_note_table(release_rows or []))

        system_fw_ws = wb.create_sheet(title="System FW")
        self._write_sheet(system_fw_ws, *build_system_fw_table(system_fw_rows or []))

        version_note_ws = wb.create_sheet(title="版本注释")
        self._write_sheet(version_note_ws, *build_version_note_table(version_note_lines or []))

        no_inventory_ws = wb.create_sheet(title="无库存部件")
        self._write_sheet(no_inventory_ws, *build_no_inventory_table(no_inventory_parts))

        wb.save(output_path)
        return output_path

    def export_change_list_reports(
        self,
        project_name: str,
        matches: Sequence[ChangeListMatch],
        date_str: Optional[str] = None,
    ) -> Dict[str, Path]:
        if not date_str:
            date_str = datetime.now().strftime("%Y%m%d")

        csv_path = self.output_dir / f"ChangeList_{project_name}_{date_str}.csv"
        txt_path = self.output_dir / f"ChangeList_{project_name}_{date_str}.txt"

        self._export_change_list_csv(csv_path, matches)
        self._export_change_list_txt(txt_path, matches)

        return {"csv": csv_path, "txt": txt_path}

    def _sanitize_testplan_rows(self, headers: List[str], rows: List[List[object]]) -> List[List[object]]:
        """
        对测试计划主表进行最终兜底清洗：
        1. 所有版本备注列统一置空
        2. HDD / SSD 的 Driver Result 强制置空
        3. HDD / SSD 若对应 OS 列有效，则 FW Result 强制为 Y
        """
        if not headers or not rows:
            return rows

        type_idx = self._find_first_index(headers, "Type")
        if type_idx is None:
            return rows

        sanitized_rows: List[List[object]] = []

        for row in rows:
            new_row = list(row)
            part_type = self._safe_upper(self._get_cell(new_row, type_idx))

            for idx, header in enumerate(headers):
                header_text = self._safe_lower(header)

                # 1) 所有版本备注列统一清空
                if self._is_version_note_header(header_text):
                    new_row[idx] = ""
                    continue

                # 2) HDD / SSD 的 Driver Result 强制清空
                if part_type in {"HDD", "SSD"} and self._is_driver_result_header(header_text):
                    new_row[idx] = ""
                    continue

                # 3) HDD / SSD 的 FW Result：只要前一个 OS 列有值，则强制 Y
                if part_type in {"HDD", "SSD"} and self._is_fw_result_header(header_text):
                    os_value = self._get_previous_os_cell_value(new_row, idx)
                    if self._is_supported_os_value(os_value):
                        new_row[idx] = "Y"
                    else:
                        new_row[idx] = ""
                    continue

            sanitized_rows.append(new_row)

        return sanitized_rows

    def _write_sheet(self, ws, headers: List[str], rows: List[List[object]]) -> None:
        ws.append(headers)
        for row in rows:
            ws.append(row)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_cells in ws.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 60)

    def _export_change_list_csv(self, path: Path, matches: Sequence[ChangeListMatch]) -> None:
        headers = [
            "release_date",
            "release_version",
            "extracted_pn",
            "exists_in_ddlist",
            "matched_ddlist_row_count",
            "status",
            "message",
        ]
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for item in matches:
                writer.writerow(
                    [
                        item.release_date,
                        item.release_version,
                        item.extracted_pn,
                        item.exists_in_ddlist,
                        item.matched_ddlist_row_count,
                        item.status,
                        item.message,
                    ]
                )

    def _export_change_list_txt(self, path: Path, matches: Sequence[ChangeListMatch]) -> None:
        with path.open("w", encoding="utf-8") as f:
            f.write("Change List 对比报告\n")
            f.write("=" * 80 + "\n")
            for item in matches:
                line = (
                    f"日期: {item.release_date} | "
                    f"版本: {item.release_version} | "
                    f"PN: {item.extracted_pn} | "
                    f"存在: {item.exists_in_ddlist} | "
                    f"匹配数: {item.matched_ddlist_row_count} | "
                    f"状态: {item.status} | "
                    f"说明: {item.message}"
                )
                f.write(line + "\n")

    # --------------------------
    # helper methods
    # --------------------------

    @staticmethod
    def _safe_upper(value: object) -> str:
        return str(value or "").strip().upper()

    @staticmethod
    def _safe_lower(value: object) -> str:
        return str(value or "").strip().lower()

    @staticmethod
    def _get_cell(row: List[object], idx: int) -> object:
        if idx < 0 or idx >= len(row):
            return ""
        return row[idx]

    @staticmethod
    def _find_first_index(headers: List[str], target: str) -> Optional[int]:
        for idx, header in enumerate(headers):
            if str(header).strip() == target:
                return idx
        return None

    @staticmethod
    def _is_fw_result_header(header_text: str) -> bool:
        """
        兼容 FW Result / FW Restult / fw result 等拼写
        """
        return "fw" in header_text and "result" in header_text

    @staticmethod
    def _is_driver_result_header(header_text: str) -> bool:
        """
        兼容 Driver Result / Driver Restult / driver result 等拼写
        """
        return "driver" in header_text and "result" in header_text

    @staticmethod
    def _is_version_note_header(header_text: str) -> bool:
        """
        版本备注列统一清空。
        兼容：
        - BIOS/BMC/FPGA/DDlist版本
        - BIOS/BMC/FPGA/Config/DDlist
        - BIOS/BMC/FPGA/DDlist 版本
        - BIOS/BMC/FPGA/Config/DDlist版本
        """
        keywords = ["bios", "bmc", "fpga", "ddlist"]
        # 只要同时带这些关键词中的核心组合，就视为备注列
        has_core = any(k in header_text for k in keywords)
        return has_core

    @staticmethod
    def _is_supported_os_value(value: object) -> bool:
        """
        判断某个 OS 支持列是否“有效”：
        - 空 / N/A / NA / None -> 视为不支持
        - inbox / 版本号 / 包名 / 链接 / Y 等 -> 视为有效
        """
        text = str(value or "").strip().lower()
        return text not in {"", "n/a", "na", "none"}

    @staticmethod
    def _get_previous_os_cell_value(row: List[object], current_idx: int) -> object:
        """
        当前假设 build_testplan_rows 的结构仍然是：
        [OS列, FW Result, Driver Result, 版本备注列]
        所以 FW Result 的前一个单元格就是对应 OS 列。
        """
        os_idx = current_idx - 1
        if os_idx < 0 or os_idx >= len(row):
            return ""
        return row[os_idx]
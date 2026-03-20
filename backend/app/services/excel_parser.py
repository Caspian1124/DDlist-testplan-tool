from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

from app.models.part_models import NormalizedPart, ReleaseNoteRow

EXPECTED_SHEETS = ["DDlist", "Release note", "System FW", "版本注释"]


@dataclass
class ParsedWorkbook:
    normalized_parts: List[NormalizedPart]
    release_note_rows: List[ReleaseNoteRow]
    system_fw_rows: List[Dict[str, Any]]
    version_note_lines: List[str]
    metadata: Dict[str, Any]


class ExcelParser:
    def __init__(self, workbook_path: str | Path):
        self.workbook_path = Path(workbook_path)
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {self.workbook_path}")
        self.wb = load_workbook(self.workbook_path, data_only=True)
        self._validate_required_sheets()

    def _validate_required_sheets(self) -> None:
        missing = [name for name in EXPECTED_SHEETS if name not in self.wb.sheetnames]
        if missing:
            raise ValueError(f"缺少必要 sheet: {missing}，当前 sheet: {self.wb.sheetnames}")

    @staticmethod
    def _clean_text(value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        return text

    @staticmethod
    def _clean_single_line_text(value: Any) -> Optional[str]:
        text = ExcelParser._clean_text(value)
        if text is None:
            return None
        return " ".join(text.split())

    @staticmethod
    def _normalize_date(value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        text = str(value).strip()
        if not text:
            return None
        return text

    @staticmethod
    def _find_header_row(ws, expected_keywords: Iterable[str], max_scan_rows: int = 10) -> int:
        expected_keywords = [k.lower() for k in expected_keywords]
        for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
            row_values = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in ws[row_idx]]
            if all(any(keyword == cell_val for cell_val in row_values) for keyword in expected_keywords):
                return row_idx
        raise ValueError(f"未找到包含关键列头 {list(expected_keywords)} 的表头行，sheet={ws.title}")

    @staticmethod
    def _row_to_dict(headers: List[str], row_values: List[Any]) -> Dict[str, Any]:
        result: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            key = header.strip() if isinstance(header, str) else str(header)
            result[key] = row_values[idx] if idx < len(row_values) else None
        return result

    def parse_ddlist_sheet(self) -> tuple[List[NormalizedPart], Dict[str, Any]]:
        ws = self.wb["DDlist"]
        header_row_idx = self._find_header_row(ws, ["Type", "PN", "L2 Description"])
        headers = [self._clean_text(cell.value) or "" for cell in ws[header_row_idx]]
        try:
            fw_idx = headers.index("FW")
            tool_idx = headers.index("Tool")
        except ValueError as exc:
            raise ValueError("DDlist sheet 缺少 FW 或 Tool 列，无法识别 OS 列范围") from exc
        os_headers = [h for h in headers[fw_idx + 1:tool_idx] if h]

        parts: List[NormalizedPart] = []
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)]
            row_dict = self._row_to_dict(headers, row_values)
            if not any(v is not None and str(v).strip() != "" for v in row_values):
                continue
            part_type = self._clean_single_line_text(row_dict.get("Type"))
            pn = self._clean_single_line_text(row_dict.get("PN"))
            l2_description = self._clean_single_line_text(row_dict.get("L2 Description"))
            if not part_type or not pn or not l2_description:
                continue
            os_raw_map = {os_col: self._clean_text(row_dict.get(os_col)) for os_col in os_headers}
            parts.append(NormalizedPart(
                part_type=part_type.upper(),
                pn=pn,
                source_tag=self._clean_single_line_text(row_dict.get("GP/CFC/HS")),
                supplier_pn=self._clean_single_line_text(row_dict.get("Supplier PN")),
                l2_description=l2_description,
                supplier=self._clean_single_line_text(row_dict.get("Supplier")),
                fw_raw=self._clean_text(row_dict.get("FW")),
                os_raw_map=os_raw_map,
                tool=self._clean_text(row_dict.get("Tool")),
                remark=self._clean_single_line_text(row_dict.get("Remark")),
                change_date=self._normalize_date(row_dict.get("change date")),
            ))
        meta = {"sheet_name": ws.title, "header_row_idx": header_row_idx, "headers": headers, "os_headers": os_headers, "total_rows": ws.max_row, "parsed_parts": len(parts)}
        return parts, meta

    def parse_release_note_sheet(self) -> tuple[List[ReleaseNoteRow], Dict[str, Any]]:
        ws = self.wb["Release note"]
        header_row_idx = self._find_header_row(ws, ["Version", "Content", "Done by"])
        headers = [self._clean_text(cell.value) or "" for cell in ws[header_row_idx]]
        rows: List[ReleaseNoteRow] = []
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)]
            row_dict = self._row_to_dict(headers, row_values)
            if not any(v is not None and str(v).strip() != "" for v in row_values):
                continue
            release_date = self._normalize_date(row_values[0] if row_values else None)
            release_version = self._clean_single_line_text(row_dict.get("Version"))
            content = self._clean_text(row_dict.get("Content"))
            done_by = self._clean_single_line_text(row_dict.get("Done by"))
            if not content:
                continue
            rows.append(ReleaseNoteRow(release_date=release_date, release_version=release_version, content=content, done_by=done_by))
        meta = {"sheet_name": ws.title, "header_row_idx": header_row_idx, "headers": headers, "total_rows": ws.max_row, "parsed_release_rows": len(rows)}
        return rows, meta

    def parse_system_fw_sheet(self) -> tuple[List[Dict[str, Any]], Dict[str, Any]]:
        ws = self.wb["System FW"]
        rows = []
        for row_idx in range(1, ws.max_row + 1):
            left = self._clean_text(ws.cell(row=row_idx, column=1).value)
            right = self._clean_text(ws.cell(row=row_idx, column=2).value)
            if left is None and right is None:
                continue
            rows.append({"row_index": row_idx, "name": left, "value": right})
        meta = {"sheet_name": ws.title, "total_rows": ws.max_row, "parsed_rows": len(rows)}
        return rows, meta

    def parse_version_note_sheet(self) -> tuple[List[str], Dict[str, Any]]:
        ws = self.wb["版本注释"]
        lines = []
        for row_idx in range(1, ws.max_row + 1):
            value = self._clean_text(ws.cell(row=row_idx, column=1).value)
            if value is not None:
                lines.append(value)
        meta = {"sheet_name": ws.title, "total_rows": ws.max_row, "parsed_lines": len(lines)}
        return lines, meta

    def parse(self) -> ParsedWorkbook:
        ddlist_parts, ddlist_meta = self.parse_ddlist_sheet()
        release_rows, release_meta = self.parse_release_note_sheet()
        system_fw_rows, system_fw_meta = self.parse_system_fw_sheet()
        version_note_lines, version_note_meta = self.parse_version_note_sheet()
        metadata = {
            "workbook_path": str(self.workbook_path),
            "sheet_names": self.wb.sheetnames,
            "ddlist": ddlist_meta,
            "release_note": release_meta,
            "system_fw": system_fw_meta,
            "version_note": version_note_meta,
        }
        return ParsedWorkbook(normalized_parts=ddlist_parts, release_note_rows=release_rows, system_fw_rows=system_fw_rows, version_note_lines=version_note_lines, metadata=metadata)


def parse_workbook(workbook_path: str | Path) -> ParsedWorkbook:
    parser = ExcelParser(workbook_path)
    return parser.parse()

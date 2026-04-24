import pandas as pd
import openpyxl
from openpyxl import load_workbook

def analyze_excel_file(filepath):
    print(f"\n{'='*60}")
    print(f"分析文件: {filepath}")
    print('='*60)
    
    wb = load_workbook(filepath)
    print(f"工作表数量: {len(wb.sheetnames)}")
    print(f"工作表名称: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n{'-'*40}")
        print(f"工作表: {sheet_name}")
        print('-'*40)
        
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, nrows=20)
        print(f"数据维度: {df.shape[0]} 行 x {df.shape[1]} 列")
        print(f"\n前20行数据预览:")
        print(df.to_string())
        
        ws = wb[sheet_name]
        print(f"\n工作表详细信息:")
        print(f"最大行数: {ws.max_row}")
        print(f"最大列数: {ws.max_column}")
        
        if ws.max_row > 0:
            print(f"\n第一行内容（可能是表头）:")
            first_row = [cell.value for cell in ws[1]]
            for i, value in enumerate(first_row, 1):
                print(f"  列{i}: {value}")

if __name__ == "__main__":
    ddlist_file = r"d:\AutoDDlistPlanTool\WR5215 G5_DDlist_ V1.0 .xlsx"
    testplan_file = r"d:\AutoDDlistPlanTool\WR5215 G5_DDlistTestplan_ V1.0.xlsx"
    
    try:
        analyze_excel_file(ddlist_file)
        analyze_excel_file(testplan_file)
    except Exception as e:
        print(f"错误: {e}")
